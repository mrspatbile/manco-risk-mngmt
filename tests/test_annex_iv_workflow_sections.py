"""Annex IV workflow section tests (MRS-198).

Covers:
- unchanged default section order for the hedge fund,
- valid standard closed-ended sections for private debt and real estate,
- valid explicit section lists for PE and infrastructure,
- no KeyError for any target fund,
- current-fund-only Excel export,
- no liquidity or redemption sections in the closed-ended notebook calls,
- investor concentration sourced from reference files, not reporting constants.
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import fund_risk_workflow.reporting.annex_iv_workflow as annex_iv_workflow
from fund_risk_workflow.config import QUARTER
from fund_risk_workflow.data.database import get_engine

PE_SECTIONS = (
    "identification", "sector_exposure", "country_exposure",
    "stage_exposure", "top5_positions", "leverage_detail",
    "performance", "aifmd_ii_disclosure",
)

INFRA_SECTIONS = (
    "identification", "asset_breakdown", "sector_breakdown",
    "country_breakdown", "top5_positions", "leverage_detail",
    "performance",
)

CLOSED_ENDED_SECTIONS = ("identification", "breakdown", "leverage_detail")

DEFAULT_SECTIONS = [
    "identification", "breakdown", "risk_measures",
    "leverage_detail", "liquidity_buckets", "liquidity_terms",
]


@pytest.fixture(scope='module')
def engine():
    return get_engine()


@pytest.fixture(autouse=True)
def no_png_export(monkeypatch):
    import fund_risk_workflow.ui.nb_utils as nb_utils
    monkeypatch.setattr(nb_utils, 'save_html_as_png',
                        lambda *a, **k: 'skipped')


def _run(engine, fund_id, sections=None):
    # output_dir must resolve inside the project root (exporter constraint);
    # 'data' is the production default used by the notebooks.
    return annex_iv_workflow.run(
        engine=engine,
        fund_id=fund_id,
        quarter=QUARTER,
        first_export_id='25',
        output_dir='data',
        sections=sections,
    )


class TestDefaultBehaviour:
    def test_hedge_fund_default_section_order_unchanged(self, engine):
        result = _run(engine, 'AIFM_HedgeFund')
        assert [name for name, _ in result['sections']] == DEFAULT_SECTIONS

    def test_hedge_fund_default_export_ids_sequential(self, engine):
        result = _run(engine, 'AIFM_HedgeFund')
        assert [eid for _, eid in result['sections']] == [
            '25', '26', '27', '28', '29', '30']


class TestClosedEndedSections:
    @pytest.mark.parametrize('fund_id',
                             ['AIFM_PrivateDebt', 'AIFM_RealEstate'])
    def test_closed_ended_runs_without_keyerror(self, engine, tmp_path, fund_id):
        result = _run(engine, fund_id,
                      sections=CLOSED_ENDED_SECTIONS)
        assert [name for name, _ in result['sections']] == list(
            CLOSED_ENDED_SECTIONS)

    @pytest.mark.parametrize('fund_id',
                             ['AIFM_PrivateDebt', 'AIFM_RealEstate'])
    def test_no_liquidity_or_redemption_sections(self, engine,
                                                 fund_id):
        result = _run(engine, fund_id,
                      sections=CLOSED_ENDED_SECTIONS)
        rendered = {name for name, _ in result['sections']}
        assert not {'liquidity_buckets', 'liquidity_terms',
                    'risk_measures'} & rendered

    @pytest.mark.parametrize('fund_id',
                             ['AIFM_PrivateDebt', 'AIFM_RealEstate'])
    def test_identification_shows_closed_ended_terms(self, engine,
                                                     fund_id):
        result = _run(engine, fund_id,
                      sections=CLOSED_ENDED_SECTIONS)
        ident = result['report']['identification']
        text = ident.astype(str).to_string()
        assert 'Closed-ended' in text


class TestExplicitPrivateAssetSections:
    def test_pe_sections_run_without_keyerror(self, engine):
        result = _run(engine, 'AIFM_PE_Buyout',
                      sections=PE_SECTIONS)
        assert [name for name, _ in result['sections']] == list(PE_SECTIONS)

    def test_infra_sections_run_without_keyerror(self, engine):
        result = _run(engine, 'AIFM_Infra_Core',
                      sections=INFRA_SECTIONS)
        assert [name for name, _ in result['sections']] == list(INFRA_SECTIONS)

    def test_unavailable_section_raises_value_error(self, engine):
        with pytest.raises(ValueError, match='unavailable'):
            _run(engine, 'AIFM_PE_Buyout',
                 sections=('identification', 'liquidity_buckets'))


class TestExport:
    def test_export_contains_current_fund_only(self, engine):
        result = _run(engine, 'AIFM_PrivateDebt',
                      sections=CLOSED_ENDED_SECTIONS)
        wb = load_workbook(Path(result['workbook_path']))
        # Fund sheets are written without the 'AIFM_' prefix
        fund_sheets = [s for s in wb.sheetnames if s != 'Summary']
        assert fund_sheets == ['PrivateDebt']


class TestInvestorConcentrationSource:
    @pytest.mark.parametrize('fund_id, n_investors',
                             [('AIFM_PrivateDebt', 8),
                              ('AIFM_RealEstate', 12)])
    def test_register_loaded_from_reference_file(self, fund_id, n_investors):
        from fund_risk_workflow.data.reference_data import load_investor_base
        path = (Path('reference_data') / 'funds' / fund_id / 'investors.json')
        raw = json.loads(path.read_text())
        base = load_investor_base(fund_id, nav_eur=100e6)
        assert len(base) == n_investors
        assert list(base['investor_id']) == [
            inv['investor_id'] for inv in raw['investors']]

    def test_register_differs_from_reporting_constants(self):
        # The approved register must be used instead of the legacy
        # hard-coded _INVESTOR_WEIGHTS reporting constants.
        from fund_risk_workflow.data.reference_data import load_investor_base
        from fund_risk_workflow.reporting.annex_iv import _INVESTOR_WEIGHTS
        base = load_investor_base('AIFM_PrivateDebt', nav_eur=100e6)
        legacy_ids = {row[0] for row in _INVESTOR_WEIGHTS['AIFM_PrivateDebt']}
        assert set(base['investor_id']) != legacy_ids
